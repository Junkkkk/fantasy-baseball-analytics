"""엑셀 리포트 생성 모듈."""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import config


# 스타일 정의
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
WIN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOSE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TIE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_report(
    matchup_df: pd.DataFrame = None,
    lineup_df: pd.DataFrame = None,
    fa_df: pd.DataFrame = None,
    strategy: list[dict] = None,
    output_dir: str = ".",
) -> str:
    """종합 엑셀 리포트를 생성한다.

    Args:
        matchup_df: 매치업 비교 DataFrame
        lineup_df: 라인업 추천 DataFrame
        fa_df: FA 추천 DataFrame
        strategy: 전략 추천 리스트
        output_dir: 출력 디렉토리

    Returns:
        생성된 파일 경로
    """
    wb = Workbook()

    # 시트 1: 매치업 현황
    if matchup_df is not None and not matchup_df.empty:
        ws = wb.active
        ws.title = "매치업 현황"
        _write_matchup_sheet(ws, matchup_df, strategy)
    else:
        ws = wb.active
        ws.title = "매치업 현황"
        ws["A1"] = "매치업 데이터 없음"

    # 시트 2: 라인업 추천
    if lineup_df is not None and not lineup_df.empty:
        ws2 = wb.create_sheet("라인업 추천")
        _write_lineup_sheet(ws2, lineup_df)

    # 시트 3: FA 추천
    if fa_df is not None and not fa_df.empty:
        ws3 = wb.create_sheet("FA 추천")
        _write_fa_sheet(ws3, fa_df)

    # 시트 4: 주간 전략
    if strategy:
        ws4 = wb.create_sheet("주간 전략")
        _write_strategy_sheet(ws4, strategy)

    # 파일 저장
    today = datetime.now().strftime("%Y%m%d")
    filename = f"fantasy_baseball_report_{today}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"리포트 생성 완료: {filepath}")
    return filepath


def _write_matchup_sheet(ws, df: pd.DataFrame, strategy: list[dict] = None):
    """매치업 현황 시트를 작성한다."""
    # 제목
    ws.merge_cells("A1:F1")
    ws["A1"] = f"📊 매치업 현황 - {datetime.now().strftime('%Y-%m-%d')}"
    ws["A1"].font = Font(bold=True, size=14)

    # 헤더
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # 데이터
    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # 결과에 따른 색상
        result_col = headers.index("result") + 1 if "result" in headers else None
        if result_col:
            result = row.get("result", "")
            fill = WIN_FILL if result == "WIN" else (LOSE_FILL if result == "LOSE" else TIE_FILL)
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = fill

    # 열 너비 조정
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15

    # 승패 요약
    summary_row = 4 + len(df) + 2
    if "result" in df.columns:
        wins = len(df[df["result"] == "WIN"])
        losses = len(df[df["result"] == "LOSE"])
        ties = len(df[df["result"] == "TIE"])
        ws.cell(row=summary_row, column=1, value="총 결과:").font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=2, value=f"{wins}승 {losses}패 {ties}무").font = Font(size=12)


def _write_lineup_sheet(ws, df: pd.DataFrame):
    """라인업 추천 시트를 작성한다."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"⚾ 라인업 추천 - {datetime.now().strftime('%Y-%m-%d')}"
    ws["A1"].font = Font(bold=True, size=14)

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # 선발/벤치 색상
        rec = row.get("recommendation", "")
        if rec == "선발":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = WIN_FILL
        elif rec == "벤치":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = PatternFill(
                    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
                )

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15


def _write_fa_sheet(ws, df: pd.DataFrame):
    """FA 추천 시트를 작성한다."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"🔍 FA 픽업 추천 - {datetime.now().strftime('%Y-%m-%d')}"
    ws["A1"].font = Font(bold=True, size=14)

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, min(len(headers) + 1, 27)):
        ws.column_dimensions[chr(64 + col_idx)].width = 15


def _write_strategy_sheet(ws, strategy: list[dict]):
    """주간 전략 시트를 작성한다."""
    ws.merge_cells("A1:C1")
    ws["A1"] = f"🎯 주간 전략 - {datetime.now().strftime('%Y-%m-%d')}"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["카테고리", "액션", "사유"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    action_colors = {
        "집중": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "수비": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "포기": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "요약": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }

    for row_idx, item in enumerate(strategy, 4):
        ws.cell(row=row_idx, column=1, value=item["category"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=item["action"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=item["reason"]).border = THIN_BORDER

        fill = action_colors.get(item["action"])
        if fill:
            for c in range(1, 4):
                ws.cell(row=row_idx, column=c).fill = fill

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
